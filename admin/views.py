from app import bcrypt_instance
from flask import (Blueprint, render_template, flash, redirect, url_for, 
                   request, jsonify, current_app, send_file)
from flask_login import login_required, current_user
from models import db, User, Service, Complaint, user_services_association
from forms import UserEditForm, ToggleUserStatusForm
from decorators import admin_required # Suponiendo que admin_required está en app.py
from sqlalchemy import func
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from datetime import datetime
admin_bp = Blueprint('admin', __name__, 
                     template_folder='../templates/admin', 
                     url_prefix='/admin')

@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    return render_template('admin_dashboard.html', active_tab='overview')



# En app.py, reemplaza la función completa por esta versión final y corregida

# En app.py, esta es la versión final de la función de exportación

@admin_bp.route('/user/<int:user_id>/exportar-excel')
@login_required
@admin_required
def exportar_excel_usuario(user_id):
    """
    Genera un archivo Excel ESTILIZADO con los detalles y servicios de un usuario.
    """
    user = User.query.get_or_404(user_id)
    user_services = user.services_contracted.all()

    if not user_services:
        flash(f"El usuario '{user.username}' no tiene servicios para exportar.", 'warning')
        return redirect(url_for('admin.user_detail', user_id=user_id))

    # 1. Preparación de datos
    datos_servicios = []
    total_amount = 0.0
    for service in user_services:
        datos_servicios.append({
            'ID Servicio': service.id, 'Nombre Servicio': service.name,
            'Tipo': service.type, 'Precio': service.price
        })
        total_amount += service.price
    df = pd.DataFrame(datos_servicios)

    # 2. Creación del archivo y escritor
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # 3. Escritura de datos
    df_cliente = pd.DataFrame([
        {'Campo': 'ID Cliente', 'Valor': user.id},
        {'Campo': 'Nombre de Usuario', 'Valor': user.username},
        {'Campo': 'Email', 'Valor': user.email},
        {'Campo': 'Fecha de Exportación', 'Valor': datetime.now().strftime('%d/%m/%Y %H:%M')}
    ])
    df_cliente.to_excel(writer, sheet_name='Info Cliente', index=False, startrow=1)
    df.to_excel(writer, sheet_name='Servicios Contratados', index=False, startrow=1)

    # 4. Acceso a objetos de openpyxl
    workbook = writer.book
    ws_cliente = workbook['Info Cliente']
    ws_servicios = workbook['Servicios Contratados']

    # 5. Definición de estilos
    font_titulo = Font(name='Calibri', size=16, bold=True)
    font_header = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    font_total = Font(name='Calibri', size=11, bold=True)
    alineacion_centrada = Alignment(horizontal='center', vertical='center')
    alineacion_derecha = Alignment(horizontal='right', vertical='center')
    relleno_header = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    formato_moneda = '$ #,##0.00'

    # 6. Aplicar estilos a 'Servicios Contratados'
    ws_servicios.insert_rows(1)
    ws_servicios.merge_cells('A1:D1')
    titulo_servicios = ws_servicios['A1']
    titulo_servicios.value = f"Reporte de Servicios de: {user.username}"
    titulo_servicios.font = font_titulo
    titulo_servicios.alignment = alineacion_centrada
    
    for cell in ws_servicios[2]:
        cell.font = font_header
        cell.fill = relleno_header
        cell.border = borde_fino
        cell.alignment = alineacion_centrada

    for row in ws_servicios.iter_rows(min_row=3, max_row=ws_servicios.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = borde_fino
        cell_precio = row[3]
        cell_precio.number_format = formato_moneda
        cell_precio.alignment = alineacion_derecha
        
    total_row_index = ws_servicios.max_row + 1
    ws_servicios[f'C{total_row_index}'] = 'TOTAL:'
    ws_servicios[f'D{total_row_index}'] = total_amount
    total_label_cell = ws_servicios[f'C{total_row_index}']
    total_value_cell = ws_servicios[f'D{total_row_index}']
    total_label_cell.font = font_total
    total_label_cell.alignment = alineacion_derecha
    total_label_cell.border = borde_fino
    total_value_cell.font = font_total
    total_value_cell.number_format = formato_moneda
    total_value_cell.border = borde_fino

    # ===== INICIO DE LA CORRECCIÓN FINAL =====
    # --- Ajustar ancho de columnas (VERSIÓN ROBUSTA) ---
    for i, column_cells in enumerate(ws_servicios.columns, 1):
        column_letter = get_column_letter(i)
        
        # Para la columna de Precio (columna 4 o 'D'), asignamos un ancho fijo
        if i == 4:
            ws_servicios.column_dimensions[column_letter].width = 15
            continue # Saltamos al siguiente ciclo del bucle

        # Para las otras columnas, usamos el cálculo automático
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_servicios.column_dimensions[column_letter].width = adjusted_width
    # ===== FIN DE LA CORRECCIÓN FINAL =====

    # 7. Aplicar estilos a 'Info Cliente'
    ws_cliente.insert_rows(1)
    ws_cliente.merge_cells('A1:B1')
    titulo_cliente = ws_cliente['A1']
    titulo_cliente.value = "Información del Cliente"
    titulo_cliente.font = font_titulo
    titulo_cliente.alignment = alineacion_centrada
    for cell in ws_cliente[2]:
        cell.font = font_header
        cell.fill = relleno_header
        cell.border = borde_fino
    for row in ws_cliente.iter_rows(min_row=3, max_row=ws_cliente.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = borde_fino
    ws_cliente.column_dimensions['A'].width = 25
    ws_cliente.column_dimensions['B'].width = 40

    # 8. Guardar y enviar el archivo
    writer.close()
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Reporte_{user.username}.xlsx'
    )


# --- NUEVA RUTA PARA LOS DATOS DEL GRÁFICO ---
@admin_bp.route('/reporte/servicios_contratados')
@login_required
@admin_required
def reporte_servicios_contratados():
    """
    Endpoint que devuelve en formato JSON la cantidad de veces que
    cada servicio ha sido contratado.
    """
    try:
        datos_grafico = db.session.query(
            Service.name, 
            func.count(user_services_association.c.user_id)
        ).join(
            user_services_association
        ).group_by(
            Service.name
        ).order_by(
            func.count(user_services_association.c.user_id).desc()
        ).all()

        labels = [row[0] for row in datos_grafico]
        data = [row[1] for row in datos_grafico]    

        return jsonify({'labels': labels, 'data': data})

    except Exception as e:
        print(f"Error al generar reporte de servicios: {e}")
        return jsonify(error="Ocurrió un error al procesar la solicitud"), 500


@admin_bp.route('/manage_users')
@login_required
@admin_required
def manage_users():
    if current_app.config['SIMULATION_MODE']:
        users = [
            {'id': 1, 'username': 'admin', 'email': 'admin@example.com', 'is_admin': True, 'is_active': True},
            {'id': 2, 'username': 'cliente', 'email': 'cliente@example.com', 'is_admin': False, 'is_active': True},
            {'id': 3, 'username': 'test', 'email': 'test@example.com', 'is_admin': False, 'is_active': False},
        ]
        toggle_form = ToggleUserStatusForm()
        return render_template('admin_user.html', users=users, active_tab='users', toggle_form=toggle_form)
    else:
        users = User.query.all()
        toggle_form = ToggleUserStatusForm()
        return render_template('admin_user.html', users=users, active_tab='users', toggle_form=toggle_form)

@admin_bp.route('/user_detail/<int:user_id>')
@login_required
@admin_required
def user_detail(user_id):
    if current_app.config['SIMULATION_MODE']:
        mock_users_data = {
            1: {'id': 1, 'username': 'admin', 'email': 'admin@example.com', 'is_admin': True, 'registration_date': '2023-01-01 10:00:00', 'last_login': '2025-05-23 18:30:00', 'is_active': True},
            2: {'id': 2, 'username': 'cliente', 'email': 'cliente@example.com', 'is_admin': False, 'registration_date': '2023-02-15 12:00:00', 'last_login': '2025-05-23 18:00:00', 'is_active': True},
            3: {'id': 3, 'username': 'test', 'email': 'test@example.com', 'is_admin': False, 'registration_date': '2023-03-20 09:00:00', 'last_login': '2025-05-23 17:00:00', 'is_active': True},
        }
        user_data_to_display = mock_users_data.get(user_id, {'id': user_id, 'username': 'Usuario Desconocido', 'email': 'unknown@example.com', 'is_admin': False, 'is_active': False})

        user_services_to_display = []
        if user_id == 2:
            user_services_to_display = [
                {'id': 101, 'name': 'Internet Fibra Óptica 300Mbps', 'type': 'Internet', 'price': 35000.00},
                {'id': 103, 'name': 'TV Cable Premium', 'type': 'TV', 'price': 25000.00},
            ]
        else:
             user_services_to_display = []

    else:
        user_data_to_display = User.query.get_or_404(user_id)
        user_services_to_display = user_data_to_display.services_contracted.all()

    toggle_form = ToggleUserStatusForm()
    return render_template('admin_detail.html', user=user_data_to_display, user_services=user_services_to_display, active_tab='users',**{'toggle_form': toggle_form})

@admin_bp.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    if current_app.config['SIMULATION_MODE']:
        flash(f'Simulando edición de usuario. (No se guardó realmente)', 'info')
        return redirect(url_for('admin.user_detail', user_id=user_id))
    else:
        user = User.query.get_or_404(user_id)
        form = UserEditForm()

        if form.validate_on_submit():
            existing_user_by_username = User.query.filter(User.username == form.username.data, User.id != user_id).first()
            existing_user_by_email = User.query.filter(User.email == form.email.data, User.id != user_id).first()

            if existing_user_by_username:
                flash('Ese nombre de usuario ya existe para otro usuario. Por favor, elige otro.', 'danger')
            elif existing_user_by_email:
                flash('Ese email ya está registrado por otro usuario. Por favor, usa otro.', 'danger')
            else:
                user.username = form.username.data
                user.email = form.email.data
                if form.password.data:
                    user.password = bcrypt_instance.generate_password_hash(form.password.data).decode('utf-8')

                if 'is_active' in form and current_user.id != user_id:
                     user.is_active = form.is_active.data
                if 'is_admin' in form:
                    user.is_admin = form.is_admin.data

                db.session.commit()
                flash('Usuario actualizado exitosamente!', 'success')
                return redirect(url_for('admin.user_detail', user_id=user.id))
        elif request.method == 'GET':
            form.username.data = user.username
            form.email.data = user.email
            if 'is_admin' in form:
                 form.is_admin.data = user.is_admin
            if 'is_active' in form:
                 form.is_active.data = user.is_active

        user_services_to_display = []
        return render_template('admin_edit_user.html', user=user, user_services=user_services_to_display, active_tab='users', form=form)

@admin_bp.route('/toggle_service_status/<int:service_id>', methods=['POST'])
@login_required
@admin_required
def toggle_service_status(service_id):
    if current_app.config['SIMULATION_MODE']:
        flash(f'Simulando cambio de estado para el servicio {service_id}. (No se guardó realmente)', 'info')
        return redirect(url_for('admin.manage_users'))
    else:
        service_item = Service.query.get_or_404(service_id)
        flash(f'Funcionalidad de cambio de estado de servicio NO IMPLEMENTADA en DB real aún.', 'warning')
        return redirect(url_for('admin.manage_users'))

@admin_bp.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    if current_app.config['SIMULATION_MODE']:
        flash(f'Simulando eliminación del usuario {user_id}. (No se eliminó realmente)', 'danger')
        return redirect(url_for('admin.manage_users'))
    else:
        user_to_delete = User.query.get_or_404(user_id)
        Complaint.query.filter_by(user_id=user_id).delete()
        db.session.delete(user_to_delete)
        db.session.commit()
        flash(f'Usuario {user_to_delete.username} eliminado permanentemente.', 'success')
        return redirect(url_for('admin.manage_users'))

@admin_bp.route('/toggle_user_active_status/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_user_active_status(user_id):
    if current_app.config['SIMULATION_MODE']:
        flash(f'Simulando cambio de estado de actividad para el usuario {user_id}. (No se guardó realmente)', 'info')
        return redirect(url_for('admin.user_detail', user_id=user_id))
    else:
        user = User.query.get_or_404(user_id)
        form = ToggleUserStatusForm()
        if form.validate_on_submit():
            if user.is_admin:
                flash('No se puede desactivar la cuenta de un administrador.', 'danger')
                return redirect(url_for('admin.user_detail', user_id=user_id))

            user.is_active = not user.is_active
            db.session.commit()
            flash(f'El usuario "{user.username}" ha sido {"activado" if user.is_active else "desactivado"} exitosamente.', 'success')
            return redirect(url_for('admin.user_detail', user_id=user.id))
        else:
            flash('Error de seguridad: Token CSRF inválido o ausente.', 'danger')
            return redirect(url_for('admin.user_detail', user_id=user_id))

@admin_bp.route('/settings')
@login_required
@admin_required
def settings():
    flash('Esta es la página de configuración del administrador (contenido estático).', 'info')
    return render_template('admin_dashboard.html', admin_section_content="<h2><i class='bi bi-gear'></i> Configuración del Sistema</h2><p>Aquí irían las opciones de configuración global del panel de administración.</p>", active_tab='settings')


@admin_bp.route('/api/user_stats')
@login_required
@admin_required
def user_stats_api():
    active_users = User.query.filter_by(is_active=True).count()
    inactive_users = User.query.filter_by(is_active=False).count()
    
    return jsonify({
        'labels': ['Usuarios Activos', 'Usuarios Inactivos'],
        'data': [active_users, inactive_users]
    })